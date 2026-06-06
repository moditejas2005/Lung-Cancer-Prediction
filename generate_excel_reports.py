import os
import glob
import logging
import warnings
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.model_selection import train_test_split, GridSearchCV, RandomizedSearchCV, StratifiedKFold
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.exceptions import ConvergenceWarning

from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.svm import SVC
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score,
    f1_score, roc_auc_score, confusion_matrix, roc_curve
)

from statsmodels.stats.outliers_influence import variance_inflation_factor
from imblearn.over_sampling import SMOTE
from imblearn.pipeline import Pipeline as ImbPipeline

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=ConvergenceWarning)
warnings.filterwarnings("ignore", category=UserWarning)

logger = logging.getLogger("ExcelReporter")

# ================================================================
# GLOBAL STYLING TOKENS (Arial font and consistent corporate palette)
# ================================================================
_side        = Side(style="thin", color="000000")
BORDER       = Border(left=_side, right=_side, top=_side, bottom=_side)
HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
ALT_FILL     = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
P1_FILL      = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
P2_FILL      = PatternFill("solid", start_color="375623", end_color="375623")
GOOD_FILL    = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
BAD_FILL     = PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7")
WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BOLD_FONT    = Font(name="Arial", bold=True, size=10)
NORMAL_FONT  = Font(name="Arial", size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=13, color="1F4E79")

def _hdr(ws, row, ncols, fill=None):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill or HEADER_FILL
        cell.font      = WHITE_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _row(ws, row, ncols, alt=False):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        if alt: cell.fill = ALT_FILL
        cell.font      = NORMAL_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

def _cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def run_vif(X_train, threshold=10.0):
    """
    Iteratively compute VIF for all features and drop variables with high VIF (> threshold).
    """
    X_work  = X_train.copy()
    dropped = []
    history = []
    it      = 1

    while True:
        vifs = []
        for i in range(X_work.shape[1]):
            try:
                v = variance_inflation_factor(X_work.values.astype(float), i)
            except:
                v = np.inf
            vifs.append(v if np.isfinite(v) else np.inf)

        df_vif = pd.DataFrame({"Feature": X_work.columns, "VIF": vifs})\
                   .sort_values("VIF", ascending=False).reset_index(drop=True)
        df_vif["Iteration"] = it
        df_vif["Status"]    = "Keep"
        history.append(df_vif.copy())

        max_vif = df_vif["VIF"].iloc[0]
        if max_vif <= threshold:
            break

        drop = df_vif.iloc[0]["Feature"]
        history[-1].loc[0, "Status"] = "DROPPED"
        dropped.append(drop)
        X_work.drop(columns=[drop], inplace=True)
        it += 1

    retained = list(X_work.columns)
    return X_work, dropped, retained, history

def evaluate(model, X, y):
    yp   = model.predict(X)
    if hasattr(model, "predict_proba"):
        ypr  = model.predict_proba(X)[:,1]
    else:
        ypr = yp
    return (
        accuracy_score(y, yp),
        precision_score(y, yp, zero_division=0),
        recall_score(y, yp, zero_division=0),
        f1_score(y, yp, zero_division=0),
        roc_auc_score(y, ypr),
        confusion_matrix(y, yp),
        roc_curve(y, ypr),
    )

def get_fi(clf, cols):
    if hasattr(clf, "feature_importances_"):
        return pd.Series(clf.feature_importances_, index=cols).sort_values(ascending=False)
    if hasattr(clf, "coef_"):
        return pd.Series(abs(clf.coef_[0]), index=cols).sort_values(ascending=False)
    return None

# ================================================================
# REPORT 1 — ModelMetrics.xlsx (8 Sheets)
# ================================================================
def generate_model_metrics_report(model_metrics, fitted_models, X_val_scaled, y_val, df_cleaned):
    logger.info("Generating ModelMetrics.xlsx (Phase 2 styled Report)...")
    wb = Workbook()
    
    # ── Sheet 1: Model_Metrics ─────────────────────────────
    ws = wb.active
    ws.title = "Model_Metrics"
    
    ws.merge_cells("A1:G1")
    ws["A1"] = "Metrics"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4B4B4B")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ["Rank", "Model", "Accuracy", "F1", "Precision", "Recall", "AUC"]
    header_colors = ["FFFF00", "FFA500", "FFFF00", "FFFF00", "FFFF00", "FFFF00", "87CEEB"]
    
    ws.append(headers)
    for col, color in enumerate(header_colors, 1):
        c = ws.cell(row=2, column=col)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name="Arial", bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        
    metrics_list = []
    for name, metrics in model_metrics.items():
        metrics_list.append([
            name,
            metrics["Accuracy"],
            metrics["F1_Score"],
            metrics["Precision"],
            metrics["Recall"],
            metrics["ROC_AUC"]
        ])
        
    metrics_sorted = sorted(metrics_list, key=lambda x: x[-1], reverse=True)
    
    row_idx = 3
    rank = 1
    for row in metrics_sorted:
        ws.append([rank] + row)
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER
            if rank % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
        row_idx += 1
        rank += 1
        
    _cw(ws, [10, 25, 14, 14, 14, 14, 14])
    
    # ── Sheet 2: Confusion_Metrices ─────────────────────────
    ws = wb.create_sheet("Confusion_Metrices")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Confusion Matrices"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0000FF")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.append(["Model", "Pred C0", "Pred C1", ""])
    for col in range(1, 5):
        c = ws.cell(row=2, column=col)
        c.fill = PatternFill("solid", fgColor="FFC000")
        c.font = Font(name="Arial", bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
        
    row = 3
    for name, model_obj in fitted_models.items():
        if name not in model_metrics:
            continue
        try:
            pred = model_obj.predict(X_val_scaled)
            cm = confusion_matrix(y_val, pred)
        except Exception:
            continue
            
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1)
        c.value = name
        c.font = Font(name="Arial", bold=True, size=12)
        c.fill = PatternFill("solid", fgColor="ADD8E6")
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        
        row += 1
        ws.cell(row=row, column=2).value = "Pred C0"
        ws.cell(row=row, column=3).value = "Pred C1"
        for col in [2, 3]:
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor="FFF2CC")
            c.font = Font(name="Arial", bold=True)
            c.alignment = Alignment(horizontal="center")
            c.border = BORDER
            
        row += 1
        ws.cell(row=row, column=1).value = "True C0"
        ws.cell(row=row, column=2).value = int(cm[0, 0])
        ws.cell(row=row, column=3).value = int(cm[0, 1])
        
        row += 1
        ws.cell(row=row, column=1).value = "True C1"
        ws.cell(row=row, column=2).value = int(cm[1, 0])
        ws.cell(row=row, column=3).value = int(cm[1, 1])
        
        for rr in [row-1, row]:
            for cc in [1, 2, 3]:
                cell = ws.cell(row=rr, column=cc)
                cell.alignment = Alignment(horizontal="center")
                cell.border = BORDER
                if (rr == row-1 and cc == 2) or (rr == row and cc == 3):
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
        row += 2
        
    _cw(ws, [25, 15, 15, 10])
    
    # ── Plots Generation for the remaining sheets ───────────
    plot_dir = "synthetic_medical_ai/data/reports/plots"
    os.makedirs(plot_dir, exist_ok=True)
    
    # Heatmap
    plt.figure(figsize=(12, 10))
    sns.heatmap(df_cleaned.select_dtypes(include=[np.number]).corr(), cmap="coolwarm", annot=False, linewidths=0.3)
    plt.title("Feature Correlation Heatmap", fontsize=14, fontweight="bold")
    plt.tight_layout()
    h_path = f"{plot_dir}/tmp_heatmap_metrics.png"
    plt.savefig(h_path, dpi=150)
    plt.close()
    
    ws = wb.create_sheet("Feature Correlation Heatmap")
    ws.add_image(XLImage(h_path), "A1")
    
    # Top 10 Feature Importances
    ws = wb.create_sheet("Top_10_Features_All_Models")
    ws.merge_cells("A1:H1")
    ws["A1"] = "Top 10 Feature Importances For All Models"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4B4B4B")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    col_idx = 0
    row_offset = 3
    columns_layout = ["A", "J"]
    
    for name, model_obj in fitted_models.items():
        # XGBoost or RF base models
        if name not in ["Optimized XGBoost", "Gradient Boosting", "Random Forest"]:
            continue
        try:
            fi = get_fi(model_obj, X_val_scaled.columns)
        except Exception:
            continue
        if fi is None:
            continue
            
        plt.figure(figsize=(6, 4))
        fi.head(10).plot(kind="barh", color="steelblue", edgecolor="black")
        plt.title(f"Top 10 Features - {name}", fontsize=12, fontweight="bold")
        plt.tight_layout()
        img_path = f"{plot_dir}/tmp_fi_{name.replace(' ', '_')}.png"
        plt.savefig(img_path, dpi=150)
        plt.close()
        
        pos = f"{columns_layout[col_idx]}{row_offset}"
        ws.add_image(XLImage(img_path), pos)
        col_idx += 1
        if col_idx >= len(columns_layout):
            col_idx = 0
            row_offset += 25
            
    # Model performance horizontal chart
    df_metrics = pd.DataFrame(metrics_list, columns=["Model", "Accuracy", "F1", "Precision", "Recall", "AUC"])
    plt.figure(figsize=(8, 5))
    df_metrics.plot(x="Model", kind="barh", edgecolor="black", linewidth=1.0)
    plt.title("Model Performance Comparison Across Metrics", fontsize=13, fontweight="bold")
    plt.xlabel("Score")
    plt.grid(axis="x", linestyle="--", alpha=0.5)
    plt.tight_layout()
    p_path = f"{plot_dir}/tmp_perf_metrics.png"
    plt.savefig(p_path, dpi=150)
    plt.close()
    
    ws = wb.create_sheet("Model_performance")
    ws.add_image(XLImage(p_path), "A1")
    
    # Subplots
    scores = ["Accuracy", "F1", "Precision", "Recall", "AUC"]
    num_models = len(df_metrics["Model"])
    raw_palette = sns.color_palette("tab10", num_models)
    palette = [matplotlib.colors.to_hex(c) for c in raw_palette]
    color_map = dict(zip(df_metrics["Model"], palette))
    
    plt.figure(figsize=(15, 3))
    for idx, score in enumerate(scores, 1):
        plt.subplot(1, 5, idx)
        bar_colors = [color_map[m] for m in df_metrics["Model"]]
        plt.barh(df_metrics["Model"], df_metrics[score], color=bar_colors, edgecolor="black", linewidth=0.6)
        plt.title(score, fontsize=11, fontweight="bold")
        plt.xlim(0, 1.05)
        plt.grid(axis="x", linestyle="--", alpha=0.3)
    plt.suptitle("Model Performance Comparison Subplots", fontsize=14, fontweight="bold", y=1.05)
    plt.tight_layout()
    p2_path = f"{plot_dir}/tmp_perf2_metrics.png"
    plt.savefig(p2_path, dpi=150)
    plt.close()
    
    ws = wb.create_sheet("Model_Performance_II")
    ws.add_image(XLImage(p2_path), "A1")
    
    # ROC Curves
    plt.figure(figsize=(8, 6))
    for name, model_obj in fitted_models.items():
        if name not in model_metrics:
            continue
        try:
            if hasattr(model_obj, "predict_proba"):
                probs = model_obj.predict_proba(X_val_scaled)[:, 1]
            else:
                probs = model_obj.predict(X_val_scaled)
            fpr, tpr, _ = roc_curve(y_val, probs)
            plt.plot(fpr, tpr, label=f"{name} (AUC = {model_metrics[name]['ROC_AUC']:.3f})", linewidth=1.8)
        except Exception:
            continue
    plt.plot([0, 1], [0, 1], "k--", lw=0.8)
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title("ROC Curves for All Models", fontsize=12, fontweight="bold")
    plt.legend(loc="lower right")
    plt.grid(alpha=0.4)
    plt.tight_layout()
    roc_path = f"{plot_dir}/tmp_roc_metrics.png"
    plt.savefig(roc_path, dpi=150)
    plt.close()
    
    ws = wb.create_sheet("ROC_Curve")
    ws.add_image(XLImage(roc_path), "A1")
    
    # Confusion Matrix Heatmaps grid
    ws = wb.create_sheet("CM_ALL")
    ws.merge_cells("A1:R1")
    ws["A1"] = "Confusion Matrices"
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4B4B4B")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    col_idx = 0
    row_offset = 3
    
    for name, model_obj in fitted_models.items():
        if name not in model_metrics:
            continue
        try:
            pred = model_obj.predict(X_val_scaled)
            cm = confusion_matrix(y_val, pred)
        except Exception:
            continue
            
        plt.figure(figsize=(4, 4))
        sns.heatmap(cm, annot=True, cmap="Blues", fmt="d", linewidths=0.5, square=True, cbar=False)
        plt.title(f"Confusion Matrix - {name}", fontsize=11, fontweight="bold")
        plt.tight_layout()
        cm_path = f"{plot_dir}/tmp_cm_{name.replace(' ', '_')}.png"
        plt.savefig(cm_path, dpi=150)
        plt.close()
        
        pos = f"{columns_layout[col_idx]}{row_offset}"
        ws.add_image(XLImage(cm_path), pos)
        
        col_idx += 1
        if col_idx >= len(columns_layout):
            col_idx = 0
            row_offset += 25
            
    # ── Sheet: Optimized_CM ────────────────────────────────
    opt_cm_path = "synthetic_medical_ai/data/reports/plots/confusion_matrix.png"
    if os.path.exists(opt_cm_path):
        ws = wb.create_sheet("Optimized_CM")
        ws.cell(1, 1, "Optimized Clinical Confusion Matrix (Calibrated Stacking)").font = TITLE_FONT
        ws.add_image(XLImage(opt_cm_path), "A3")
        
    # Save ModelMetrics.xlsx
    try:
        wb.save("synthetic_medical_ai/data/reports/ModelMetrics.xlsx")
        logger.info("Successfully saved ModelMetrics.xlsx!")
    except PermissionError:
        alt_path = "synthetic_medical_ai/data/reports/ModelMetrics_new.xlsx"
        logger.warning(f"Permission denied to write ModelMetrics.xlsx (locked by another process). Saving to {alt_path} instead.")
        wb.save(alt_path)

# ================================================================
# REPORT 2 — Final_Evaluation_Report.xlsx (13 Sheets)
# ================================================================
def generate_combined_evaluation_report(df_cleaned, test_mode=False, fitted_models=None, model_metrics=None, X_val_scaled=None, y_val=None):
    logger.info("Executing Two-Phase ML Pipeline (VIF vs PCA)...")
    plot_dir = "synthetic_medical_ai/data/reports/plots"
    os.makedirs(plot_dir, exist_ok=True)
    
    # Engineer features and clean
    from synthetic_medical_ai.preprocessing.feature_engineering import engineer_features
    df_feat = engineer_features(df_cleaned)
    
    X = df_feat.drop("Diagnosis", axis=1)
    y = df_feat["Diagnosis"]
    
    # 80/20 Stratified train-test split
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, stratify=y, test_size=0.2, random_state=42
    )
    
    cv = StratifiedKFold(n_splits=3 if test_mode else 5, shuffle=True, random_state=42)
    
    # ── Phase 1: VIF Feature Selection ─────────────────────
    # Run VIF on continuous numeric columns only
    num_cols = X_train.select_dtypes(include=[np.number]).columns
    X_train_vif_input = X_train[num_cols].copy()
    X_train_vif, dropped_cols, retained_cols, vif_history = run_vif(X_train_vif_input, threshold=10.0)
    
    X_test_vif = X_test[retained_cols]
    
    # Models grid search configuration
    # Small hyper-parameter grids for fast local run
    phase1_models = {
        "Logistic Regression": GridSearchCV(
            ImbPipeline([
                ("smote",  SMOTE(random_state=42)),
                ("scaler", StandardScaler()),
                ("clf", LogisticRegression(max_iter=1000, class_weight="balanced")),
            ]),
            {"clf__C": [0.1, 1.0] if not test_mode else [1.0]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "Random Forest": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", RandomForestClassifier(random_state=42, class_weight="balanced")),
            ]),
            {"clf__n_estimators": [200] if not test_mode else [100], "clf__max_depth": [10, None] if not test_mode else [10]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "SVM": GridSearchCV(
            ImbPipeline([
                ("smote",  SMOTE(random_state=42)),
                ("scaler", StandardScaler()),
                ("clf", SVC(probability=True, class_weight="balanced")),
            ]),
            {"clf__C": [1.0]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "Gradient Boosting": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", GradientBoostingClassifier(random_state=42)),
            ]),
            {"clf__n_estimators": [100], "clf__learning_rate": [0.1]},
            cv=cv, scoring="recall", n_jobs=-1
        )
    }
    
    # Train Phase 1
    m1, cm1, roc1, fi1, hp1 = [], {}, {}, {}, []
    for name, gs in phase1_models.items():
        logger.info(f"Training Phase 1 VIF model: {name}...")
        gs.fit(X_train_vif, y_train)
        best = gs.best_estimator_
        acc, pre, rec, f1, auc, cm, roc = evaluate(best, X_test_vif, y_test)
        m1.append([name, acc, pre, rec, f1, auc])
        cm1[name]  = cm
        roc1[name] = roc
        fi1[name]  = get_fi(best.named_steps["clf"], retained_cols)
        hp1.append({"model": name, "params": gs.best_params_, "cv_recall": gs.best_score_})
        
    df_m1 = pd.DataFrame(m1, columns=["Model", "Accuracy", "Precision", "Recall", "F1", "AUC"])
    
    # ── Phase 2: PCA Feature Transformation ──────────────────
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train[num_cols])
    X_test_scaled  = scaler.transform(X_test[num_cols])
    
    pca = PCA(n_components=0.95, random_state=42)
    pca.fit(X_train_scaled)
    n_pca = pca.n_components_
    
    X_train_pca = pca.transform(X_train_scaled)
    X_test_pca  = pca.transform(X_test_scaled)
    
    pca_cols = [f"PC{i+1}" for i in range(n_pca)]
    
    phase2_models = {
        "Logistic Regression": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", LogisticRegression(max_iter=1000, class_weight="balanced")),
            ]),
            {"clf__C": [0.1, 1.0] if not test_mode else [1.0]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "Random Forest": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", RandomForestClassifier(random_state=42, class_weight="balanced")),
            ]),
            {"clf__n_estimators": [200] if not test_mode else [100], "clf__max_depth": [10, None] if not test_mode else [10]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "SVM": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", SVC(probability=True, class_weight="balanced")),
            ]),
            {"clf__C": [1.0]},
            cv=cv, scoring="recall", n_jobs=-1
        ),
        "Gradient Boosting": GridSearchCV(
            ImbPipeline([
                ("smote", SMOTE(random_state=42)),
                ("clf", GradientBoostingClassifier(random_state=42)),
            ]),
            {"clf__n_estimators": [100], "clf__learning_rate": [0.1]},
            cv=cv, scoring="recall", n_jobs=-1
        )
    }
    
    # Train Phase 2
    m2, cm2, roc2, fi2, hp2 = [], {}, {}, {}, []
    for name, gs in phase2_models.items():
        logger.info(f"Training Phase 2 PCA model: {name}...")
        gs.fit(X_train_pca, y_train)
        best = gs.best_estimator_
        acc, pre, rec, f1, auc, cm, roc = evaluate(best, X_test_pca, y_test)
        m2.append([name, acc, pre, rec, f1, auc])
        cm2[name]  = cm
        roc2[name] = roc
        fi2[name]  = get_fi(best.named_steps["clf"], pca_cols)
        hp2.append({"model": name, "params": gs.best_params_, "cv_recall": gs.best_score_})
        
    df_m2 = pd.DataFrame(m2, columns=["Model", "Accuracy", "Precision", "Recall", "F1", "AUC"])
    
    # Combined Rankings
    all_rows = [(r + ["Phase 1"]) for r in m1] + [(r + ["Phase 2"]) for r in m2]
    df_all   = pd.DataFrame(all_rows, columns=["Model","Accuracy","Precision","Recall","F1","AUC","Phase"])
    df_all_s = df_all.sort_values("Recall", ascending=False).reset_index(drop=True)
    df_all_s.index += 1
    
    # ── Generate Final_Evaluation_Report.xlsx Workbook ──────────
    logger.info("Writing 13 sheets for Final_Evaluation_Report.xlsx...")
    wb = Workbook()
    wb.remove(wb.active)
    
    # ── Sheet 1: Summary ──────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.cell(1, 1, "ML PIPELINE — COMBINED RESULTS SUMMARY").font = Font(name="Arial", bold=True, size=15, color="1F4E79")
    ws.merge_cells("A1:H1")
    
    info = [
        ("Dataset",          f"{X.shape[0]} rows × {X.shape[1]} features"),
        ("Train / Test",     f"{X_train.shape[0]} / {X_test.shape[0]} (80/20 stratified)"),
        ("CV Strategy",      "StratifiedKFold (5-fold)"),
        ("Imbalance",        "SMOTE on train folds only"),
        ("Optimize Metric",  "Recall"),
        ("Phase 1 Method",   f"VIF ≤ 10 → {len(retained_cols)} features retained, {len(dropped_cols)} dropped"),
        ("Phase 2 Method",   f"PCA 95% variance → {n_pca} components from {len(num_cols)} features"),
    ]
    r = 3
    for label, val in info:
        ws.cell(r, 1, label).font = BOLD_FONT
        ws.cell(r, 2, val).font   = NORMAL_FONT
        r += 1
        
    r += 1
    ws.cell(r, 1, "COMBINED MODEL METRICS (sorted by Recall)").font = TITLE_FONT
    r += 1
    headers = ["Rank", "Phase", "Model", "Accuracy", "Precision", "Recall", "F1", "AUC"]
    for c, h in enumerate(headers, 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 8)
    _cw(ws, [8, 12, 28, 14, 14, 14, 14, 14])
    r += 1
    
    for rank, row in df_all_s.iterrows():
        ws.cell(r, 1, rank)
        ws.cell(r, 2, row["Phase"])
        ws.cell(r, 3, row["Model"])
        for c, col in enumerate(["Accuracy", "Precision", "Recall", "F1", "AUC"], 4):
            ws.cell(r, c, round(row[col], 4))
        if rank == 1:
            for c in range(1, 9):
                ws.cell(r, c).fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
        _row(ws, r, 8, alt=(rank%2==0))
        ws.cell(r, 2).font = BOLD_FONT
        r += 1
        
    # ── Sheet 2: Phase1_VIF_History ─────────────────────────
    ws = wb.create_sheet("Phase1_VIF_History")
    ws.cell(1, 1, "Phase 1 — VIF Iterative Elimination History").font = TITLE_FONT
    ws.merge_cells("A1:E1")
    _cw(ws, [8, 30, 14, 14, 14])
    
    r = 3
    for df_it in vif_history:
        it_num = df_it["Iteration"].iloc[0]
        ws.cell(r, 1, f"Iteration {it_num}").font = BOLD_FONT
        r += 1
        for c, h in enumerate(["#", "Feature", "VIF", "Status"], 1):
            ws.cell(r, c, h)
        _hdr(ws, r, 4, fill=P1_FILL)
        r += 1
        for idx, row_data in df_it.iterrows():
            ws.cell(r, 1, idx+1)
            ws.cell(r, 2, row_data["Feature"])
            vif_val = round(float(row_data["VIF"]), 3) if np.isfinite(row_data["VIF"]) else "∞"
            ws.cell(r, 3, vif_val)
            ws.cell(r, 4, row_data["Status"])
            _row(ws, r, 4, alt=(idx%2==1))
            if row_data["Status"] == "DROPPED":
                for c in range(1, 5):
                    ws.cell(r, c).fill = BAD_FILL
            r += 1
        r += 1
        
    ws.cell(r, 1, "Retained Features").font = BOLD_FONT; r += 1
    for i, f in enumerate(retained_cols):
        ws.cell(r, 1, i+1)
        ws.cell(r, 2, f)
        ws.cell(r, 2).fill = GOOD_FILL
        r += 1
    r += 1
    ws.cell(r, 1, "Dropped Features").font = BOLD_FONT; r += 1
    for i, f in enumerate(dropped_cols):
        ws.cell(r, 1, i+1)
        ws.cell(r, 2, f)
        ws.cell(r, 2).fill = BAD_FILL
        r += 1
        
    # ── Sheet 3: Phase1_Model_Metrics ────────────────────────
    ws = wb.create_sheet("Phase1_Model_Metrics")
    ws.cell(1, 1, "Phase 1 — Model Performance (VIF-selected features)").font = TITLE_FONT
    _cw(ws, [8, 28, 14, 14, 14, 14, 14])
    r = 3
    for c, h in enumerate(["Rank", "Model", "Accuracy", "Precision", "Recall", "F1", "AUC"], 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 7, fill=P1_FILL); r += 1
    for rank, row in enumerate(sorted(m1, key=lambda x: x[3], reverse=True), 1):
        ws.cell(r, 1, rank)
        for c, v in enumerate(row, 2):
            ws.cell(r, c, v if isinstance(v, str) else round(float(v), 4))
        _row(ws, r, 7, alt=(rank%2==0)); r += 1
        
    # ── Sheet 4: Phase1_Hyperparameters ──────────────────────
    ws = wb.create_sheet("Phase1_Hyperparameters")
    ws.cell(1, 1, "Phase 1 — Best Hyperparameters").font = TITLE_FONT
    _cw(ws, [28, 38, 22, 18])
    r = 3
    for c, h in enumerate(["Model", "Parameter", "Best Value", "CV Best Recall"], 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 4, fill=P1_FILL); r += 1
    for rec in hp1:
        first = True
        for k, v in rec["params"].items():
            ws.cell(r, 1, rec["model"] if first else "")
            ws.cell(r, 2, k)
            ws.cell(r, 3, str(v))
            ws.cell(r, 4, round(rec["cv_recall"], 4) if first else "")
            _row(ws, r, 4, alt=(r%2==0)); first = False; r += 1
        r += 1
        
    # ── Sheet 5: Phase1_Confusion_Matrices ────────────────────
    ws = wb.create_sheet("Phase1_Confusion_Matrices")
    ws.cell(1, 1, "Phase 1 — Confusion Matrices").font = TITLE_FONT
    _cw(ws, [22, 14, 14])
    r = 3
    for name, cm in cm1.items():
        ws.cell(r, 1, name).font = BOLD_FONT; r += 1
        for c, h in enumerate(["", "Pred 0", "Pred 1"], 1):
            ws.cell(r, c, h)
        _hdr(ws, r, 3, fill=P1_FILL); r += 1
        for lbl, row_data in zip(["Actual 0", "Actual 1"], cm):
            ws.cell(r, 1, lbl).font = BOLD_FONT
            for c, v in enumerate(row_data, 2):
                ws.cell(r, c, int(v))
            _row(ws, r, 3); r += 1
        r += 2
        
    # ── Sheet 6: Phase1_Charts ──────────────────────────────
    ws = wb.create_sheet("Phase1_Charts")
    ws.cell(1, 1, "Phase 1 — Charts").font = TITLE_FONT
    ri = 3
    
    # 1. Performance bars
    fig, axes = plt.subplots(1, 5, figsize=(18, 4))
    for ax, met in zip(axes, ["Accuracy", "Precision", "Recall", "F1", "AUC"]):
        ax.barh(df_m1["Model"], df_m1[met], color="#2E75B6")
        ax.set_xlim(0, 1.05)
        ax.set_title(met, fontweight="bold")
    plt.suptitle("Phase 1 — Performance per Metric", fontsize=13, fontweight="bold")
    plt.tight_layout()
    p1_p = f"{plot_dir}/tmp_p1_perf.png"
    plt.savefig(p1_p, dpi=150)
    plt.close()
    ws.add_image(XLImage(p1_p), f"A{ri}"); ri += 22
    
    # 2. ROC curves
    plt.figure(figsize=(7, 4))
    for name, (fpr, tpr, _) in roc1.items():
        plt.plot(fpr, tpr, label=name)
    plt.plot([0, 1], [0, 1], "k--", lw=0.8)
    plt.xlabel("FPR")
    plt.ylabel("TPR")
    plt.title("Phase 1 — ROC Curves")
    plt.legend()
    plt.grid(alpha=0.3)
    plt.tight_layout()
    p1_r = f"{plot_dir}/tmp_p1_roc.png"
    plt.savefig(p1_r, dpi=150)
    plt.close()
    ws.add_image(XLImage(p1_r), f"A{ri}"); ri += 22
    
    # 3. CM Heatmaps
    fig, axes = plt.subplots(1, len(cm1), figsize=(4*len(cm1), 3.5))
    if len(cm1) == 1:
        axes = [axes]
    for ax, (name, cm) in zip(axes, cm1.items()):
        sns.heatmap(cm, annot=True, fmt="d", cmap="Blues", ax=ax, cbar=False)
        ax.set_title(name, fontsize=10)
        ax.set_xlabel("Predicted")
        ax.set_ylabel("Actual")
    plt.suptitle("Phase 1 — Confusion Matrices", fontsize=12, fontweight="bold")
    plt.tight_layout()
    p1_cm = f"{plot_dir}/tmp_p1_cm.png"
    plt.savefig(p1_cm, dpi=150)
    plt.close()
    ws.add_image(XLImage(p1_cm), f"A{ri}"); ri += 22
    
    # 4. Correlation heatmap
    plt.figure(figsize=(10, 8))
    sns.heatmap(X_train_vif.corr(), cmap="coolwarm", annot=False, linewidths=0.3)
    plt.title("Phase 1 — Feature Correlation Heatmap", fontsize=13, fontweight="bold")
    plt.tight_layout()
    p1_corr = f"{plot_dir}/tmp_p1_corr.png"
    plt.savefig(p1_corr, dpi=150)
    plt.close()
    ws.add_image(XLImage(p1_corr), f"A{ri}")
    
    # ── Sheet 7: Phase2_PCA_Info ────────────────────────────
    ws = wb.create_sheet("Phase2_PCA_Info")
    ws.cell(1, 1, "Phase 2 — PCA Information").font = Font(name="Arial", bold=True, size=13, color="375623")
    _cw(ws, [16, 20, 20])
    r = 3
    ws.cell(r, 1, "Component")
    ws.cell(r, 2, "Explained Var %")
    ws.cell(r, 3, "Cumulative %")
    _hdr(ws, r, 3, fill=P2_FILL); r += 1
    
    cumev = np.cumsum(pca.explained_variance_ratio_)
    for i, (ev, ce) in enumerate(zip(pca.explained_variance_ratio_, cumev)):
        ws.cell(r, 1, f"PC{i+1}")
        ws.cell(r, 2, round(ev*100, 3))
        ws.cell(r, 3, round(ce*100, 3))
        _row(ws, r, 3, alt=(i%2==1)); r += 1
    r += 2
    
    # Scree Plot
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.bar(range(1, len(ev)+1) if hasattr(ev, "__len__") else [1], [ev*100] if not hasattr(ev, "__len__") else ev*100, alpha=0.7, label="Individual")
    ax.plot(range(1, len(cumev)+1) if hasattr(cumev, "__len__") else [1], [cumev*100] if not hasattr(cumev, "__len__") else cumev*100, "r-o", ms=4, label="Cumulative")
    ax.axhline(95, color="gray", ls="--", lw=0.8, label="95% threshold")
    ax.set_xlabel("Component")
    ax.set_ylabel("Explained Variance %")
    ax.set_title("Phase 2 — PCA Scree Plot")
    ax.legend()
    plt.tight_layout()
    p2_scree = f"{plot_dir}/tmp_p2_scree.png"
    plt.savefig(p2_scree, dpi=150)
    plt.close()
    ws.add_image(XLImage(p2_scree), f"A{r}")
    
    # ── Sheet 8: Phase2_PCA_Loadings ─────────────────────────
    ws = wb.create_sheet("Phase2_PCA_Loadings")
    ws.cell(1, 1, "Phase 2 — PCA Loadings (feature contribution to each PC)").font = Font(name="Arial", bold=True, size=13, color="375623")
    
    n_show = min(n_pca, 15)
    show_pc_cols = [f"PC{i+1}" for i in range(n_show)]
    loadings = pd.DataFrame(pca.components_[:n_show].T, index=num_cols, columns=show_pc_cols)
    
    # Loadings heatmap plot
    plt.figure(figsize=(10, 8))
    sns.heatmap(loadings, cmap="RdBu_r", center=0, annot=True, fmt=".2f", linewidths=0.3)
    plt.title("Phase 2 — PCA Loadings Heatmap", fontsize=13, fontweight="bold")
    plt.tight_layout()
    p2_load = f"{plot_dir}/tmp_p2_load.png"
    plt.savefig(p2_load, dpi=150)
    plt.close()
    ws.add_image(XLImage(p2_load), "A3")
    
    # Table layout cells below image
    row_offset = len(num_cols)*3 + 25
    _cw(ws, [28] + [10]*n_show)
    ws.cell(row_offset, 1, "Feature")
    for j, pc in enumerate(show_pc_cols, 2):
        ws.cell(row_offset, j, pc)
    _hdr(ws, row_offset, n_show+1, fill=P2_FILL)
    
    for ri2, (feat, row_data) in enumerate(loadings.iterrows()):
        rr = row_offset + 1 + ri2
        ws.cell(rr, 1, feat).font = BOLD_FONT
        for j, v in enumerate(row_data, 2):
            cell = ws.cell(rr, j, round(float(v), 4))
            cell.font = NORMAL_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            if abs(v) > 0.3:
                cell.fill = PatternFill("solid", start_color="FFD7D7", end_color="FFD7D7") if v > 0 \
                       else PatternFill("solid", start_color="D7E8FF", end_color="D7E8FF")
        ws.cell(rr, 1).border = BORDER
        _row(ws, rr, n_show+1, alt=(ri2%2==0))
        ws.cell(rr, 1).font = BOLD_FONT
        
    # ── Sheet 9: Phase2_Model_Metrics ────────────────────────
    ws = wb.create_sheet("Phase2_Model_Metrics")
    ws.cell(1, 1, "Phase 2 — Model Performance (PCA-transformed features)").font = Font(name="Arial", bold=True, size=13, color="375623")
    _cw(ws, [8, 28, 14, 14, 14, 14, 14])
    r = 3
    for c, h in enumerate(["Rank", "Model", "Accuracy", "Precision", "Recall", "F1", "AUC"], 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 7, fill=P2_FILL); r += 1
    for rank, row in enumerate(sorted(m2, key=lambda x: x[3], reverse=True), 1):
        ws.cell(r, 1, rank)
        for c, v in enumerate(row, 2):
            ws.cell(r, c, v if isinstance(v, str) else round(float(v), 4))
        _row(ws, r, 7, alt=(rank%2==0)); r += 1
        
    # ── Sheet 10: Phase2_Hyperparameters ─────────────────────
    ws = wb.create_sheet("Phase2_Hyperparameters")
    ws.cell(1, 1, "Phase 2 — Best Hyperparameters").font = Font(name="Arial", bold=True, size=13, color="375623")
    _cw(ws, [28, 38, 22, 18])
    r = 3
    for c, h in enumerate(["Model", "Parameter", "Best Value", "CV Best Recall"], 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 4, fill=P2_FILL); r += 1
    for rec in hp2:
        first = True
        for k, v in rec["params"].items():
            ws.cell(r, 1, rec["model"] if first else "")
            ws.cell(r, 2, k)
            ws.cell(r, 3, str(v))
            ws.cell(r, 4, round(rec["cv_recall"], 4) if first else "")
            _row(ws, r, 4, alt=(r%2==0)); first = False; r += 1
        r += 1
        
    # ── Sheet 11: Phase2_Confusion_Matrices ──────────────────
    ws = wb.create_sheet("Phase2_Confusion_Matrices")
    ws.cell(1, 1, "Phase 2 — Confusion Matrices").font = Font(name="Arial", bold=True, size=13, color="375623")
    _cw(ws, [22, 14, 14]); r = 3
    for name, cm in cm2.items():
        ws.cell(r, 1, name).font = BOLD_FONT; r += 1
        for c, h in enumerate(["", "Pred 0", "Pred 1"], 1):
            ws.cell(r, c, h)
        _hdr(ws, r, 3, fill=P2_FILL); r += 1
        for lbl, row_data in zip(["Actual 0", "Actual 1"], cm):
            ws.cell(r, 1, lbl).font = BOLD_FONT
            for c, v in enumerate(row_data, 2):
                ws.cell(r, c, int(v))
            _row(ws, r, 3); r += 1
        r += 2
        
    # ── Sheet 12: Phase2_Charts ─────────────────────────────
    ws = wb.create_sheet("Phase2_Charts")
    ws.cell(1, 1, "Phase 2 — Charts").font = Font(name="Arial", bold=True, size=13, color="375623")
    ri = 3
    
    # 1. Performance bars
    fig, axes = plt.subplots(1, 5, figsize=(18, 4))
    for ax, met in zip(axes, ["Accuracy", "Precision", "Recall", "F1", "AUC"]):
        ax.barh(df_m2["Model"], df_m2[met], color="#375623")
        ax.set_xlim(0, 1.05)
        ax.set_title(met, fontweight="bold")
    plt.suptitle("Phase 2 — Performance per Metric", fontsize=13, fontweight="bold")
    plt.tight_layout()
    p2_p = f"{plot_dir}/tmp_p2_perf.png"
    plt.savefig(p2_p, dpi=150)
    plt.close()
    ws.add_image(XLImage(p2_p), f"A{ri}"); ri += 22
    
    # 2. ROC curves
    plt.figure(figsize=(7, 4))
    for name, (fpr, tpr, _) in roc2.items():
        plt.plot(fpr, tpr, label=name)
    plt.plot([0, 1], [0, 1], "k--", lw=0.8)
    plt.xlabel("FPR")
    plt.ylabel("TPR")
    plt.title("Phase 2 — ROC Curves")
    plt.legend()
    plt.grid(alpha=0.3)
    plt.tight_layout()
    p2_r = f"{plot_dir}/tmp_p2_roc.png"
    plt.savefig(p2_r, dpi=150)
    plt.close()
    ws.add_image(XLImage(p2_r), f"A{ri}"); ri += 22
    
    # 3. CM heatmaps
    fig, axes = plt.subplots(1, len(cm2), figsize=(4*len(cm2), 3.5))
    if len(cm2) == 1:
        axes = [axes]
    for ax, (name, cm) in zip(axes, cm2.items()):
        sns.heatmap(cm, annot=True, fmt="d", cmap="Greens", ax=ax, cbar=False)
        ax.set_title(name, fontsize=10)
        ax.set_xlabel("Predicted")
        ax.set_ylabel("Actual")
    plt.suptitle("Phase 2 — Confusion Matrices", fontsize=12, fontweight="bold")
    plt.tight_layout()
    p2_cm = f"{plot_dir}/tmp_p2_cm.png"
    plt.savefig(p2_cm, dpi=150)
    plt.close()
    ws.add_image(XLImage(p2_cm), f"A{ri}")
    
    # ── Sheet 13: Phase_Comparison ─────────────────────────
    ws = wb.create_sheet("Phase_Comparison")
    ws.cell(1, 1, "Phase 1 vs Phase 2 — Full Comparison").font = TITLE_FONT
    _cw(ws, [8, 12, 28, 14, 14, 14, 14, 14])
    r = 3
    for c, h in enumerate(["Rank", "Phase", "Model", "Accuracy", "Precision", "Recall", "F1", "AUC"], 1):
        ws.cell(r, c, h)
    _hdr(ws, r, 8); r += 1
    
    for rank, row in df_all_s.iterrows():
        ws.cell(r, 1, rank)
        ws.cell(r, 2, row["Phase"])
        ws.cell(r, 3, row["Model"])
        for c, col in enumerate(["Accuracy", "Precision", "Recall", "F1", "AUC"], 4):
            ws.cell(r, c, round(row[col], 4))
        fill = P1_FILL if row["Phase"] == "Phase 1" else P2_FILL
        ws.cell(r, 2).fill = fill
        ws.cell(r, 2).font = WHITE_FONT
        _row(ws, r, 8, alt=(rank%2==0))
        ws.cell(r, 2).fill = fill
        ws.cell(r, 2).font = WHITE_FONT
        r += 1
    r += 2
    
    # Side-by-side metrics chart
    fig, axes = plt.subplots(1, 5, figsize=(20, 5))
    metrics_list = ["Accuracy", "Precision", "Recall", "F1", "AUC"]
    for ax, met in zip(axes, metrics_list):
        vals1 = [v[metrics_list.index(met)] for name, *v in m1]
        vals2 = [v[metrics_list.index(met)] for name, *v in m2]
        names1 = [n for n, *v in m1]
        names2 = [n for n, *v in m2]
        
        ax.barh([f"P1:{n}" for n in names1], vals1, color="#2E75B6", alpha=0.8)
        ax.barh([f"P2:{n}" for n in names2], vals2, color="#375623", alpha=0.8)
        ax.set_xlim(0, 1.05)
        ax.set_title(met, fontweight="bold")
        ax.axvline(0.5, color="gray", ls="--", lw=0.5)
    plt.suptitle("Phase 1 (Blue) vs Phase 2 (Green) — Performance Comparison", fontsize=14, fontweight="bold")
    plt.tight_layout()
    compare_p = f"{plot_dir}/tmp_compare.png"
    plt.savefig(compare_p, dpi=150)
    plt.close()
    ws.add_image(XLImage(compare_p), f"A{r}")
    
    # ── Sheet: CM_ALL ──────────────────────────────────────
    if fitted_models is not None and model_metrics is not None and X_val_scaled is not None and y_val is not None:
        ws = wb.create_sheet("CM_ALL")
        ws.merge_cells("A1:R1")
        ws["A1"] = "Confusion Matrices"
        ws["A1"].font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4B4B4B")
        ws["A1"].alignment = Alignment(horizontal="center")
        
        columns_layout = ["A", "J"]
        col_idx = 0
        row_offset = 3
        
        for name, model_obj in fitted_models.items():
            if name not in model_metrics:
                continue
            try:
                pred = model_obj.predict(X_val_scaled)
                cm = confusion_matrix(y_val, pred)
            except Exception:
                continue
                
            plt.figure(figsize=(4, 4))
            sns.heatmap(cm, annot=True, cmap="Blues", fmt="d", linewidths=0.5, square=True, cbar=False)
            plt.title(f"Confusion Matrix - {name}", fontsize=11, fontweight="bold")
            plt.tight_layout()
            cm_path = f"{plot_dir}/tmp_cm_{name.replace(' ', '_')}.png"
            plt.savefig(cm_path, dpi=150)
            plt.close()
            
            pos = f"{columns_layout[col_idx]}{row_offset}"
            ws.add_image(XLImage(cm_path), pos)
            
            col_idx += 1
            if col_idx >= len(columns_layout):
                col_idx = 0
                row_offset += 25

    # ── Sheet: Optimized_CM ────────────────────────────────
    opt_cm_path = "synthetic_medical_ai/data/reports/plots/confusion_matrix.png"
    if os.path.exists(opt_cm_path):
        ws = wb.create_sheet("Optimized_CM")
        ws.cell(1, 1, "Optimized Clinical Confusion Matrix (Calibrated Stacking)").font = TITLE_FONT
        ws.add_image(XLImage(opt_cm_path), "A3")
        
    # Save Final_Evaluation_Report.xlsx
    try:
        wb.save("synthetic_medical_ai/data/reports/Final_Evaluation_Report.xlsx")
        logger.info("Successfully saved Final_Evaluation_Report.xlsx!")
    except PermissionError:
        alt_path = "synthetic_medical_ai/data/reports/Final_Evaluation_Report_new.xlsx"
        logger.warning(f"Permission denied to write Final_Evaluation_Report.xlsx (locked by another process). Saving to {alt_path} instead.")
        wb.save(alt_path)
    
    # Cleanup temporary charts
    for pattern in ["tmp_*.png", f"{plot_dir}/tmp_*.png"]:
        for f in glob.glob(pattern):
            try:
                os.remove(f)
            except Exception:
                pass
    logger.info("Excel reports generation cleanup completed successfully.")
